"""
pg_to_excel.py
==============
Exports every table in a PostgreSQL schema to an Excel workbook.
Each sheet = one table, with:
  Row 1  : table name + table description (merged cells)
  Row 2  : column headers
  Row 3+ : column metadata

Requirements:
    pip install psycopg2-binary openpyxl python-dotenv

.env file (same directory):
    DB_HOST=localhost
    DB_PORT=5432
    DB_NAME=mydb
    DB_USER=myuser
    DB_PASSWORD=mypassword
    DB_SCHEMA=public          # optional, defaults to 'public'
    OUTPUT_FILE=schema.xlsx   # optional
"""

import os
import re
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────

load_dotenv()

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 5432)),
    "dbname":   os.getenv("DB_NAME"),
    "user":     os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
}
SCHEMA      = os.getenv("DB_SCHEMA", "public")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "schema_documentation.xlsx")

# Max characters for a valid sheet name
MAX_SHEET_NAME = 31

# ── Styles ────────────────────────────────────────────────────────────────────

COLOUR = {
    "table_banner_bg":  "1F3864",   # dark navy
    "table_banner_fg":  "FFFFFF",
    "header_bg":        "2E75B6",   # medium blue
    "header_fg":        "FFFFFF",
    "pk_bg":            "FFF2CC",   # light yellow  — primary key rows
    "alt_bg":           "EBF3FB",   # alternating row tint
    "border":           "B0C4DE",
}

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _font(bold=False, colour="000000", size=11):
    return Font(bold=bold, color=colour, size=size)

def _border():
    side = Side(style="thin", color=COLOUR["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Database helpers ──────────────────────────────────────────────────────────

TABLES_SQL = """
SELECT
    t.table_name,
    obj_description(pc.oid, 'pg_class') AS table_description
FROM information_schema.tables t
JOIN pg_class pc ON pc.relname = t.table_name
JOIN pg_namespace pn ON pn.oid = pc.relnamespace
    AND pn.nspname = t.table_schema
WHERE t.table_schema = %s
  AND t.table_type = 'BASE TABLE'
ORDER BY t.table_name;
"""

COLUMNS_SQL = """
SELECT
    c.column_name,
    c.udt_name                              AS column_type,
    c.character_maximum_length,
    c.numeric_precision,
    c.numeric_scale,
    c.is_nullable,
    c.column_default,
    col_description(pc.oid, c.ordinal_position) AS column_description,
    -- constraints (comma-separated)
    COALESCE(
        STRING_AGG(DISTINCT tc.constraint_type, ', '
            ORDER BY tc.constraint_type),
        ''
    )                                        AS constraints,
    -- indexed?
    CASE WHEN COUNT(DISTINCT pi.indexrelid) > 0 THEN 'Yes' ELSE 'No' END AS is_indexed
FROM information_schema.columns c
JOIN pg_class pc ON pc.relname = c.table_name
JOIN pg_namespace pn ON pn.oid = pc.relnamespace
    AND pn.nspname = c.table_schema
LEFT JOIN information_schema.key_column_usage kcu
    ON  kcu.table_schema  = c.table_schema
    AND kcu.table_name    = c.table_name
    AND kcu.column_name   = c.column_name
LEFT JOIN information_schema.table_constraints tc
    ON  tc.constraint_name  = kcu.constraint_name
    AND tc.table_schema     = c.table_schema
    AND tc.table_name       = c.table_name
LEFT JOIN pg_attribute pa
    ON  pa.attrelid = pc.oid
    AND pa.attname  = c.column_name
    AND pa.attnum   > 0
LEFT JOIN pg_index pi ON pi.indrelid = pc.oid
    AND pa.attnum = ANY(pi.indkey)
WHERE c.table_schema = %s
  AND c.table_name   = %s
GROUP BY
    c.column_name, c.udt_name, c.character_maximum_length,
    c.numeric_precision, c.numeric_scale, c.is_nullable,
    c.column_default, c.ordinal_position, pc.oid
ORDER BY c.ordinal_position;
"""

def connect():
    return psycopg2.connect(**DB_CONFIG)

def fetch_tables(cur):
    cur.execute(TABLES_SQL, (SCHEMA,))
    return cur.fetchall()   # [(table_name, table_description), ...]

def fetch_columns(cur, table_name):
    cur.execute(COLUMNS_SQL, (SCHEMA, table_name))
    return cur.fetchall()

def format_type(udt_name, char_max, num_prec, num_scale):
    """Build a human-readable type string, e.g. varchar(255), numeric(10,2)."""
    t = udt_name
    if char_max:
        t = f"{t}({char_max})"
    elif num_prec and num_scale:
        t = f"{t}({num_prec},{num_scale})"
    elif num_prec:
        t = f"{t}({num_prec})"
    return t

# ── Excel builder ─────────────────────────────────────────────────────────────

HEADERS = [
    "Column Name",
    "Column Type",
    "Nullable",
    "Default Value",
    "Constraint",
    "Index",
    "Column Description",
]
COL_COUNT = len(HEADERS)

def safe_sheet_name(name, existing):
    """Truncate to 31 chars and deduplicate."""
    base = re.sub(r'[\\/*?:\[\]]', '_', name)[:MAX_SHEET_NAME]
    candidate, n = base, 1
    while candidate in existing:
        suffix = f"_{n}"
        candidate = base[:MAX_SHEET_NAME - len(suffix)] + suffix
        n += 1
    return candidate

def write_table_sheet(wb, table_name, table_desc, columns):
    sheet_names = [ws.title for ws in wb.worksheets]
    ws = wb.create_sheet(title=safe_sheet_name(table_name, sheet_names))

    # ── Row 1: Table banner ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
    banner_cell = ws.cell(row=1, column=1)
    desc_text   = table_desc or "(no description)"
    banner_cell.value     = f"Table: {table_name}    |    {desc_text}"
    banner_cell.font      = _font(bold=True, colour=COLOUR["table_banner_fg"], size=12)
    banner_cell.fill      = _fill(COLOUR["table_banner_bg"])
    banner_cell.alignment = _center()
    ws.row_dimensions[1].height = 24

    # ── Row 2: Column headers ────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = _font(bold=True, colour=COLOUR["header_fg"])
        cell.fill      = _fill(COLOUR["header_bg"])
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[2].height = 20

    # ── Rows 3+: Column data ─────────────────────────────────────────────────
    for row_offset, col in enumerate(columns):
        (col_name, udt, char_max, num_prec, num_scale,
         nullable, default, col_desc, constraints, is_indexed) = col

        is_pk    = "PRIMARY KEY" in (constraints or "")
        row_num  = row_offset + 3
        row_fill = _fill(COLOUR["pk_bg"]) if is_pk else (
                   _fill(COLOUR["alt_bg"]) if row_offset % 2 == 0 else None)

        values = [
            col_name,
            format_type(udt, char_max, num_prec, num_scale),
            "Yes" if nullable == "YES" else "No",
            default or "",
            constraints or "",
            is_indexed,
            col_desc or "",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = _left()
            cell.border    = _border()
            if row_fill:
                cell.fill = row_fill
            if col_name and col_idx == 1 and is_pk:
                cell.font = _font(bold=True)

        ws.row_dimensions[row_num].height = 18

    # ── Auto-fit column widths ───────────────────────────────────────────────
    for col_idx in range(1, COL_COUNT + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(HEADERS[col_idx - 1])
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    return ws

# ── Summary sheet ─────────────────────────────────────────────────────────────

def write_summary_sheet(wb, table_info):
    ws = wb.create_sheet(title="📋 Summary", index=0)

    # Banner
    ws.merge_cells("A1:C1")
    cell = ws.cell(row=1, column=1, value=f"Schema: {SCHEMA}  —  Database Documentation")
    cell.font      = _font(bold=True, colour=COLOUR["table_banner_fg"], size=13)
    cell.fill      = _fill(COLOUR["table_banner_bg"])
    cell.alignment = _center()
    ws.row_dimensions[1].height = 26

    # Headers
    for col_idx, hdr in enumerate(["Table Name", "Description", "Column Count"], start=1):
        c = ws.cell(row=2, column=col_idx, value=hdr)
        c.font = _font(bold=True, colour=COLOUR["header_fg"])
        c.fill = _fill(COLOUR["header_bg"])
        c.alignment = _center()
        c.border = _border()

    for row_idx, (tname, tdesc, col_count) in enumerate(table_info, start=3):
        fill = _fill(COLOUR["alt_bg"]) if row_idx % 2 == 0 else None
        for col_idx, val in enumerate([tname, tdesc or "", col_count], start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = _left()
            c.border = _border()
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Connecting to '{DB_CONFIG['dbname']}' on {DB_CONFIG['host']}…")
    conn = connect()
    cur  = conn.cursor()

    tables = fetch_tables(cur)
    if not tables:
        print(f"No tables found in schema '{SCHEMA}'.")
        return

    print(f"Found {len(tables)} table(s). Building workbook…")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    summary_rows = []

    for table_name, table_desc in tables:
        columns = fetch_columns(cur, table_name)
        write_table_sheet(wb, table_name, table_desc, columns)
        summary_rows.append((table_name, table_desc, len(columns)))
        print(f"  ✔  {table_name:<40} ({len(columns)} columns)")

    write_summary_sheet(wb, summary_rows)

    cur.close()
    conn.close()

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved → {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
